"""
Match names between two sheets of an Excel file.
- Sheet1 names go in column A
- Best match from Sheet2 goes in column B
- Alternative options (if no perfect match) go in column C
"""
import openpyxl
from Levenshtein import ratio
import re

INPUT_FILE = 'שיבוץ מול מקור.xlsx'
OUTPUT_FILE = 'שיבוץ מול מקור - התאמה.xlsx'

EXACT_THRESHOLD = 0.85  # Above this = confident match
OPTION_THRESHOLD = 0.55  # Above this = show as option
HIGH_CONFIDENCE = 0.99   # Above this = don't show options

# Hebrew nickname/short-name equivalences (bidirectional)
NICKNAME_GROUPS = [
    {'אבי', 'אברהם'},
    {'איציק', 'יצחק'},
    {'ויקי', 'ויקטוריה'},
    {'דן', 'דני', 'דניאל'},
    {'מוטי', 'מרדכי'},
    {'שלומי', 'שלמה'},
    {'יוסי', 'יוסף'},
    {'משה', 'מושי', 'מושון'},
    {'דוד', 'דודי', 'דודו'},
    {'חיים', 'חיימי', 'חיימקה'},
    {'יעקב', 'יעקי', 'קובי'},
    {'שמעון', 'שמעי', 'שימי'},
    {'בני', 'בנימין', 'בנצי'},
    {'אלי', 'אליהו'},
    {'רפי', 'רפאל'},
    {'גבי', 'גבריאל'},
    {'מיכי', 'מיכאל'},
    {'אריק', 'אריה'},
    {'ישי', 'ישראל'},
    {'שמואל', 'שמוליק', 'שמוליק'},
    {'נתן', 'נתנאל'},
    {'עמי', 'עמוס'},
    {'צחי', 'יצחק'},
    {'ריקי', 'רבקה'},
    {'שרה', 'שרי', 'שריתה'},
    {'רחל', 'רחלי', 'רחלה'},
    {'מירי', 'מרים'},
    {'חני', 'חנה'},
    {'אסתי', 'אסתר'},
    {'שושי', 'שושנה'},
    {'ציפי', 'ציפורה'},
    {'רני', 'רנה', 'רונן'},
]

# Build lookup: word -> set of equivalent words
NICKNAME_LOOKUP = {}
for group in NICKNAME_GROUPS:
    for name in group:
        if name not in NICKNAME_LOOKUP:
            NICKNAME_LOOKUP[name] = set()
        NICKNAME_LOOKUP[name].update(group - {name})


def are_nicknames(word1: str, word2: str) -> bool:
    """Check if two words are known nickname equivalents."""
    if word1 in NICKNAME_LOOKUP and word2 in NICKNAME_LOOKUP[word1]:
        return True
    return False


def normalize(name: str) -> str:
    """Remove extra whitespace and strip."""
    if not name:
        return ''
    return re.sub(r'\s+', ' ', name.strip())

def name_parts(name: str) -> set:
    """Split name into parts (words)."""
    return set(normalize(name).split())

def match_score(name1: str, name2: str) -> float:
    """
    Score how well two names match.
    Handles reversed order (first/last name swap) and fuzzy similarity.
    """
    n1 = normalize(name1)
    n2 = normalize(name2)

    if not n1 or not n2:
        return 0.0

    # Exact match
    if n1 == n2:
        return 1.0

    # Same parts in different order (e.g. "יניב פנחס" vs "פנחס יניב")
    parts1 = name_parts(n1)
    parts2 = name_parts(n2)
    if parts1 == parts2:
        return 0.99

    # Check if parts match when considering nicknames
    # e.g. "אבי כהן" vs "אברהם כהן"
    if len(parts1) > 0 and len(parts2) > 0 and len(parts1) == len(parts2):
        parts1_list = list(parts1)
        parts2_list = list(parts2)
        # Try to match all parts (exact or nickname)
        nickname_matched = _try_nickname_match(parts1_list, parts2_list)
        if nickname_matched:
            return 0.97

    # Check if all parts of one name appear in the other (subset match)
    # e.g. "אבו אלעז מוחמד" vs "אבו אלעז מוחמד חסן"
    if parts1.issubset(parts2) or parts2.issubset(parts1):
        return 0.95

    # Fuzzy match on individual parts (handles typos + nicknames)
    if len(parts1) > 0 and len(parts2) > 0:
        parts1_list = list(parts1)
        parts2_list = list(parts2)
        total_part_score = 0
        matched_parts2 = set()

        for p1 in parts1_list:
            best_part_score = 0
            best_idx = -1
            for idx, p2 in enumerate(parts2_list):
                if idx in matched_parts2:
                    continue
                # Check nickname first
                if p1 == p2:
                    s = 1.0
                elif are_nicknames(p1, p2):
                    s = 0.95
                else:
                    s = ratio(p1, p2)
                if s > best_part_score:
                    best_part_score = s
                    best_idx = idx
            if best_idx >= 0:
                matched_parts2.add(best_idx)
                total_part_score += best_part_score

        parts_ratio = total_part_score / max(len(parts1_list), len(parts2_list))

        # Also compute overall string similarity
        str_ratio = ratio(n1, n2)

        return max(parts_ratio, str_ratio)

    return ratio(n1, n2)


def _try_nickname_match(parts1: list, parts2: list) -> bool:
    """Check if all parts match via exact or nickname equivalence (any order)."""
    used = set()
    for p1 in parts1:
        found = False
        for idx, p2 in enumerate(parts2):
            if idx in used:
                continue
            if p1 == p2 or are_nicknames(p1, p2):
                used.add(idx)
                found = True
                break
        if not found:
            return False
    return len(used) == len(parts2)


def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    sheets = wb.sheetnames

    ws1 = wb[sheets[0]]
    ws2 = wb[sheets[1]]

    # Read names from both sheets (column A)
    names1 = []
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val and str(val).strip():
            names1.append(normalize(str(val)))

    names2 = []
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val and str(val).strip():
            names2.append(normalize(str(val)))

    print(f"Sheet1 ({sheets[0]}): {len(names1)} names")
    print(f"Sheet2 ({sheets[1]}): {len(names2)} names")

    # Create output workbook
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = 'התאמה'
    out_ws.sheet_properties.sheetFormatPr = None

    # Set RTL
    out_ws.sheet_view.rightToLeft = True

    # Headers
    out_ws.cell(1, 1, f'{sheets[0]}')
    out_ws.cell(1, 2, f'{sheets[1]} - התאמה')
    out_ws.cell(1, 3, 'ציון')
    out_ws.cell(1, 4, 'אפשרויות נוספות')

    # Bold headers
    from openpyxl.styles import Font
    bold = Font(bold=True, size=12)
    for col in range(1, 5):
        out_ws.cell(1, col).font = bold

    # Match each name from sheet1 to sheet2
    used_matches = set()  # Track which sheet2 names were used as primary match

    results = []
    for name1 in names1:
        scores = []
        for name2 in names2:
            s = match_score(name1, name2)
            scores.append((s, name2))

        scores.sort(key=lambda x: -x[0])

        best_score, best_match = scores[0] if scores else (0, '')

        # Get options (other decent matches)
        options = []
        for s, n in scores[1:6]:  # Top 5 alternatives
            if s >= OPTION_THRESHOLD:
                options.append(f"{n} ({s:.0%})")

        if best_score >= EXACT_THRESHOLD:
            results.append((name1, best_match, best_score, options))
        else:
            # No confident match
            all_options = []
            for s, n in scores[:5]:
                if s >= OPTION_THRESHOLD:
                    all_options.append(f"{n} ({s:.0%})")
            results.append((name1, '', 0, all_options))

    # Write results
    row_num = 2
    matched_count = 0
    unmatched_count = 0

    for name1, match, score, options in results:
        out_ws.cell(row_num, 1, name1)
        if match:
            out_ws.cell(row_num, 2, match)
            out_ws.cell(row_num, 3, f'{score:.0%}')
            matched_count += 1
        else:
            out_ws.cell(row_num, 2, 'לא נמצאה התאמה')
            unmatched_count += 1

        if options and score < HIGH_CONFIDENCE:
            out_ws.cell(row_num, 4, ' | '.join(options))

        row_num += 1

    # Auto-width columns
    for col in range(1, 5):
        max_len = 0
        for row in range(1, row_num):
            val = out_ws.cell(row, col).value
            if val:
                max_len = max(max_len, len(str(val)))
        out_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 2, 80)

    # Also add a sheet for names in Sheet2 that weren't matched
    matched_names2 = {r[1] for r in results if r[1]}
    unmatched_from_2 = [n for n in names2 if n not in matched_names2]

    if unmatched_from_2:
        ws_unmatched = out_wb.create_sheet(f'לא הותאמו מ-{sheets[1]}')
        ws_unmatched.sheet_view.rightToLeft = True
        ws_unmatched.cell(1, 1, f'שמות מ-{sheets[1]} ללא התאמה')
        ws_unmatched.cell(1, 1).font = bold
        for i, name in enumerate(unmatched_from_2, start=2):
            ws_unmatched.cell(i, 1, name)
        ws_unmatched.column_dimensions['A'].width = 30

    out_wb.save(OUTPUT_FILE)
    print(f"\nResults saved to: {OUTPUT_FILE}")
    print(f"Matched: {matched_count}, Unmatched: {unmatched_count}")
    if unmatched_from_2:
        print(f"Names in {sheets[1]} with no match from {sheets[0]}: {len(unmatched_from_2)}")


if __name__ == '__main__':
    main()
